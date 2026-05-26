"""
Carga y normalización de datos desde archivos Excel municipales.
"""
from __future__ import annotations

import re
from pathlib import Path

import openpyxl
import pandas as pd

_EXPECTED_COLUMNS = ["TITULAR", "COTITULAR", "LOTE", "DIRECCION", "ESTADO", "OBSERVACIONES"]
_SHEET_PATTERN = re.compile(r"^Manzana\s+\d+$", re.IGNORECASE)

# Columnas distintivas del formato plano (hoja única)
_PLANO_REQUIRED = {"TITULARES", "ESCRITURADO"}
_PLANO_ALL = ["TITULARES", "MANZANA", "LOTE", "DIRECCION", "ESCRITURADO", "OBSERVACION"]


def _detect_plano_sheet(wb) -> str | None:
    """Devuelve el nombre de la primera hoja que tiene columnas TITULARES + ESCRITURADO, o None."""
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if first_row is None:
            continue
        header = {str(c).strip().upper() for c in first_row if c is not None}
        if _PLANO_REQUIRED.issubset(header):
            return sheet_name
    return None


def _cargar_formato_manzanas(wb, label: str, manzana_sheets: list[str]) -> pd.DataFrame:
    """Carga el formato multi-hoja 'Manzana N' (formato HUMITOS)."""
    frames: list[pd.DataFrame] = []

    for sheet_name in manzana_sheets:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            raise ValueError(
                f"La hoja '{sheet_name}' está vacía y no tiene fila de encabezado."
            )

        header_row = rows[0]
        data_rows = rows[1:]

        header = [str(c).strip() if c is not None else "" for c in header_row]

        missing = [col for col in _EXPECTED_COLUMNS if col not in header]
        if missing:
            raise ValueError(
                f"La hoja '{sheet_name}' no tiene las columnas requeridas: {missing}. "
                f"Columnas encontradas: {header}"
            )

        col_idx = {col: header.index(col) for col in _EXPECTED_COLUMNS}

        records = []
        for row in data_rows:
            titular = row[col_idx["TITULAR"]]
            titular_s = str(titular).strip() if titular is not None else ""
            if not titular_s or titular_s.lower() == "total":
                continue

            estado = row[col_idx["ESTADO"]]
            if estado is not None:
                estado = str(estado).strip().upper()

            records.append({
                "TITULAR": titular,
                "COTITULAR": row[col_idx["COTITULAR"]],
                "LOTE": row[col_idx["LOTE"]],
                "DIRECCION": row[col_idx["DIRECCION"]],
                "ESTADO": estado,
                "OBSERVACIONES": row[col_idx["OBSERVACIONES"]],
                "MANZANA": sheet_name,
            })

        frames.append(pd.DataFrame(records, columns=_EXPECTED_COLUMNS + ["MANZANA"]))

    if not frames:
        raise ValueError(
            "No se encontraron datos válidos en ninguna hoja 'Manzana N'."
        )

    return pd.concat(frames, ignore_index=True)


def _cargar_formato_plano(wb, label: str, sheet_name: str) -> pd.DataFrame:
    """Carga el formato de hoja única con columnas TITULARES, ESCRITURADO, MANZANA, etc."""
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        raise ValueError(f"La hoja '{sheet_name}' está vacía.")

    header_row = rows[0]
    data_rows = rows[1:]

    header = [str(c).strip().upper() if c is not None else "" for c in header_row]

    missing = [col for col in _PLANO_ALL if col not in header]
    if missing:
        raise ValueError(
            f"La hoja '{sheet_name}' no tiene columnas requeridas: {missing}. "
            f"Columnas encontradas: {header}"
        )

    col_idx = {col: header.index(col) for col in _PLANO_ALL}
    legajo_idx = header.index("LEGAJO") if "LEGAJO" in header else None

    records = []
    for row in data_rows:
        # Filtrar fila "Total"
        first_cell = row[0]
        if first_cell is not None and str(first_cell).strip().lower() == "total":
            continue

        titulares_raw = row[col_idx["TITULARES"]]
        if titulares_raw is None or str(titulares_raw).strip() == "":
            continue

        # Separar titular y cotitular
        titulares_str = str(titulares_raw).strip()
        if " - " in titulares_str:
            parts = titulares_str.split(" - ", 1)
            titular = parts[0].strip()
            cotitular = parts[1].strip() if parts[1].strip() else None
        else:
            titular = titulares_str
            cotitular = None

        # Mapear ESCRITURADO → ESTADO
        escriturado_raw = row[col_idx["ESCRITURADO"]]
        if escriturado_raw is not None and str(escriturado_raw).strip().upper() == "SI":
            estado = "ESCRITURADO"
        else:
            estado = "NO ESCRITURADO"

        # Formatear MANZANA
        manzana_raw = row[col_idx["MANZANA"]]
        manzana = f"Manzana {str(manzana_raw).strip()}" if manzana_raw is not None else None

        record = {
            "TITULAR": titular,
            "COTITULAR": cotitular,
            "LOTE": row[col_idx["LOTE"]],
            "DIRECCION": row[col_idx["DIRECCION"]],
            "ESTADO": estado,
            "OBSERVACIONES": row[col_idx["OBSERVACION"]],
            "MANZANA": manzana,
        }
        if legajo_idx is not None:
            record["LEGAJO"] = row[legajo_idx]

        records.append(record)

    if not records:
        raise ValueError(f"No se encontraron datos válidos en la hoja '{sheet_name}'.")

    output_cols = _EXPECTED_COLUMNS + ["MANZANA"]
    if legajo_idx is not None:
        output_cols = output_cols + ["LEGAJO"]

    return pd.DataFrame(records, columns=output_cols)


def cargar_excel(path: str | Path | object) -> pd.DataFrame:
    """
    Lee un archivo Excel municipal y devuelve un DataFrame unificado.

    Detecta automáticamente dos formatos:
    - Multi-hoja "Manzana N" (HUMITOS): columnas TITULAR, COTITULAR, LOTE, DIRECCION, ESTADO, OBSERVACIONES
    - Hoja única plana (Martinez Estrada): columnas TITULARES, MANZANA, LOTE, DIRECCION, ESCRITURADO, OBSERVACION

    Retorna
    -------
    pd.DataFrame
        Columnas: TITULAR, COTITULAR, LOTE, DIRECCION, ESTADO, OBSERVACIONES, MANZANA
        (más LEGAJO si está presente en el archivo).

    Lanza
    -----
    ValueError
        Si el archivo no puede leerse o no coincide con ningún formato conocido.
    """
    source = path if hasattr(path, "read") else Path(path)
    label = getattr(path, "name", str(path))

    try:
        wb = openpyxl.load_workbook(source, data_only=True)
    except Exception as exc:
        raise ValueError(
            f"No se pudo leer el archivo '{label}': {exc}"
        ) from exc

    # Formato multi-hoja "Manzana N"
    manzana_sheets = [name for name in wb.sheetnames if _SHEET_PATTERN.match(name)]
    if manzana_sheets:
        return _cargar_formato_manzanas(wb, label, manzana_sheets)

    # Formato hoja única con TITULARES + ESCRITURADO
    plano_sheet = _detect_plano_sheet(wb)
    if plano_sheet is not None:
        return _cargar_formato_plano(wb, label, plano_sheet)

    raise ValueError(
        f"No se reconoce el formato del archivo '{label}'. "
        f"Se esperan hojas 'Manzana N' o una hoja con columnas TITULARES y ESCRITURADO. "
        f"Hojas disponibles: {wb.sheetnames}"
    )
